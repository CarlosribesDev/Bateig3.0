from openpyxl import Workbook
from openpyxl.styles import Font,Alignment
from openpyxl.styles.borders import Border, Side
import openpyxl
import datetime

def get_day(date):
    months = {
        1 : 'Enero',
        2 : 'Febrero',
        3 : 'Marzo',
        4 : 'Abril',
        5 : 'Mayo',
        6 : 'Junio',
        7 : 'Julio',
        8 : 'Agosto',
        9 : 'Septimebre',
        10 : 'Octubre' ,
        11 : 'Noviembre',
        12 : 'Diciembre'
    }
    week_days = {
        0 : 'Lunes',
        1 : 'Martes',
        2 : 'Miercoles',
        3 : 'Jueves',
        4 : 'Viernes',
        5 : 'Sabado',
        6 : 'Domingo'
    }
    data_date = date.split("/")

    day = int(data_date[0])
    month = int(data_date[1])
    year = int(data_date[2])

    week_day = datetime.datetime(year,month,day)

    week = datetime.date(year, month, day).isocalendar().week

    return week_days[week_day.weekday()],week,str(day),months[month],str(year)



def export_to_excel(report):

    week_day,week_num,day,month,year = get_day(report.date)

    name = '{} {} {} semana {}.xlsx'.format(report.name,month,year,week_num)

    try:
        book = openpyxl.load_workbook(name)
       
    except:
        book = Workbook()
        
    sheet = book.create_sheet(week_day)

    #formato de la hoja
    
    sheet['A1'] = day + ' de ' + month + ' de ' + year
    sheet['D1'] = 'LOSAS' 
    sheet['M1'] = 'TABLAS'
    sheet['A2'] = 'Fecha'
    sheet['B2'] = 'Turno'
    sheet['C2'] = 'Operario'
    sheet['D2'] = 'L'
    sheet['E2'] = 'A'
    sheet['F2'] = 'G'
    sheet['G2'] = 'Piezas'
    sheet['H2'] = 'M2'
    sheet['I2'] = 'Material'
    sheet['J2'] = 'Faena'
    sheet['K2'] = 'Total Losas'
    sheet['L2'] = 'Total M2 losas'
    sheet['M2'] = 'Material Tabla'
    sheet['N2'] = 'Grueso tabla'
    sheet['O2'] = 'Tablas'
    sheet['P2'] = 'M2 Tabla'
    sheet['Q2'] = 'Total Tablas'
    sheet['R2'] = 'Total M2 Tablas'
    

    write_data(report,sheet) 
    styles(sheet)

    book.save(name)

    


def write_data(report,sheet):

    row = 3
      
    turn = '{}:{} - {}:{}'.format(report.start_hour,report.start_minute,report.end_hour,report.end_minute)
    
    #mostrar cortes   
    for measure in report.measures_list:

        #datos medida
        sheet['A{}'.format(row)] = report.date
        sheet['B{}'.format(row)] = turn
        sheet['C{}'.format(row)] = report.name
        sheet['D{}'.format(row)] = measure.length
        sheet['E{}'.format(row)] = measure.width
        sheet['F{}'.format(row)] = measure.thick
        sheet['G{}'.format(row)] = measure.amount
        sheet['H{}'.format(row)] = '= ({} * {} / 10000) * {}'.format('D{}'.format(row),'E{}'.format(row),'G{}'.format(row))
        sheet['H{}'.format(row)].number_format = '0.00'          
        sheet['I{}'.format(row)] = measure.material

        row +=1
    
    row = 3
    #mostrar tablas
    for measure in report.measures_slab_list:
        sheet['M{}'.format(row)] = measure.material
        sheet['N{}'.format(row)] = measure.thick
        sheet['O{}'.format(row)] = measure.amount
        sheet['P{}'.format(row)] = round(measure.m2,2)
        row += 1
    
    num_rows = 0
    for row in sheet.rows:
        num_rows +=1
    #mostrar totales
    sheet['K3'] = '=SUM(G3:G{})'.format(num_rows)
    # sheet['L3'] = round(report.total_m2,2)
    sheet['L3'] = '=SUM(H3:H{})'.format(num_rows)
    sheet['L3'].number_format = '0.00'
    sheet['Q3'] = '=SUM(O3:O{})'.format(num_rows)    
    sheet['R3'] = '=SUM(P3:P{})'.format(num_rows)    
  
  
def styles(sheet):

    #agrupar celdas
    sheet.merge_cells('A1:C1')
    sheet.merge_cells('D1:L1')
    sheet.merge_cells('M1:R1')
   

    #alinear todas las celdas
    alignment = Alignment(horizontal = 'center',vertical='center') 
    
    for row in sheet.rows:
        for cell in row:
            if cell.value:
                cell.alignment = alignment
    
    #separa por fecha losas tablas
    for cell in sheet ['D:D']:
        cell.border = Border(left=Side(style='thick'))
    
    for cell in sheet ['M:M']:
        cell.border = Border(left=Side(style='thick'))
    #color celdas de titulos
    my_color = openpyxl.styles.colors.Color(rgb='00FFFF99')
    my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_color)
    my_border = Border(left=Side(style='thick'), 
                     right=Side(style='thick'), 
                     top=Side(style='thick'), 
                     bottom=Side(style='thick'))
    my_font = Font(bold = True)

    for cell in sheet['1:1']:
        cell.font = my_font
        cell.border = my_border
    
    #estilo titulos columnas
    for cell in sheet['2:2']:
        cell.font = my_font
        cell.fill = my_fill
        cell.border = my_border
    
    
    #aumenta ancho de las celdas
    colums = ['A','B','C','I','J','K','L','M','N','O','P','Q','R']
    
    for colum in colums:
       sheet.column_dimensions[colum].width = 15

    #centrar texto observaciones
    # sheet['N3'].alignment = Alignment(vertical='top',horizontal = 'left')
    
    
    
   
